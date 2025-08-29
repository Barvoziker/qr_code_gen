#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import qrcode.image.svg
import os
import sys
import base64
import platform
from pathlib import Path
import openpyxl



def create_vcard(prenom, nom, profession, societe, mobile, pro, email, adresse, site_web):
    """
    Crée une vCard ultra-optimisée pour réduire la complexité du QR code
    """
    # Format ultra-compact - version 2.1 et syntaxe minimale
    vcard_lines = ["BEGIN:VCARD", "VERSION:3.0"]
    
    # Nom (obligatoire) - format standard
    full_name = f"{prenom} {nom}".strip()
    if full_name:
        vcard_lines.append(f"FN:{full_name}")
        vcard_lines.append(f"N:{nom};{prenom}")
    
    # Téléphones - format court sans TYPE
    if mobile and str(mobile).strip() and str(mobile).strip().lower() != 'nan':
        vcard_lines.append(f"TEL:{mobile}")
    
    if pro and str(pro).strip() and str(pro).strip().lower() != 'nan':
        vcard_lines.append(f"TEL:{pro}")
    
    # Email - format court
    if email and str(email).strip() and str(email).strip().lower() != 'nan':
        vcard_lines.append(f"EMAIL:{email}")
    
    # Société - format court
    if societe and str(societe).strip() and str(societe).strip().lower() != 'nan':
        vcard_lines.append(f"ORG:{societe}")
    
    # Profession - format court
    if profession and str(profession).strip() and str(profession).strip().lower() != 'nan':
        vcard_lines.append(f"TITLE:{profession}")
    
    # Site web - format court sans https://
    if site_web and str(site_web).strip() and str(site_web).strip().lower() != 'nan':
        url = str(site_web).strip()
        # Enlever le protocole pour économiser des caractères
        if url.startswith('https://'):
            url = url[8:]
        elif url.startswith('http://'):
            url = url[7:]
        vcard_lines.append(f"URL:{url}")
    
    vcard_lines.append("END:VCARD")
    
    return "\n".join(vcard_lines)


def generate_qr_svg(data, filename):
    """
    Génère un QR code en format SVG de 100x100 pixels
    """
    # Encoder les données en UTF-8 pour éviter les problèmes d'encodage
    if isinstance(data, str):
        data = data.encode('utf-8').decode('utf-8')
    
    # Configuration pour QR code 100x100 pixels
    qr = qrcode.QRCode(
        version=1,  # Taille minimale (21x21 modules)
        error_correction=qrcode.constants.ERROR_CORRECT_L,  # Correction d'erreur faible
        box_size=4,  # 4px par module (21*4 = 84px + bordures = ~100px)
        border=2,   # Bordure de 2 modules (8px)
    )
    
    qr.add_data(data)
    qr.make(fit=True)
    
    factory = qrcode.image.svg.SvgPathImage
    img = qr.make_image(image_factory=factory)
    
    # Déterminer le chemin approprié pour les QR codes
    # Si l'application est compilée en .exe, utiliser le dossier de l'utilisateur
    if getattr(sys, 'frozen', False):
        # L'application est compilée avec PyInstaller
        # Utiliser le dossier Documents de l'utilisateur
        
        if platform.system() == 'Windows':
            # Sur Windows, utiliser le dossier Documents
            user_docs = os.path.join(os.path.expanduser('~'), 'Documents')
            output_dir = Path(user_docs) / "Sudalys_QR_Codes"
        else:
            # Sur d'autres systèmes, utiliser le dossier home
            output_dir = Path(os.path.expanduser('~')) / "Sudalys_QR_Codes"
    else:
        # En mode développement, utiliser le dossier local
        output_dir = Path("qr_codes")
    
    # Créer le dossier de sortie s'il n'existe pas
    output_dir.mkdir(exist_ok=True, parents=True)
    
    # Sauvegarder le fichier SVG avec encodage UTF-8
    filepath = output_dir / f"{filename}.svg"
    with open(filepath, 'wb') as f:
        img.save(f)
    
    # Corriger l'encodage du fichier SVG généré
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            svg_content = f.read()
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(svg_content)
    except UnicodeDecodeError:
        # Si erreur d'encodage, essayer avec latin-1 puis sauver en UTF-8
        with open(filepath, 'r', encoding='latin-1') as f:
            svg_content = f.read()
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(svg_content)
    
    return filepath


def is_empty_cell(cell):
    """
    Vérifie si une cellule est vide ou contient 'nan'
    """
    if cell is None:
        return True
    cell_value = str(cell).strip()
    return cell_value == '' or cell_value.lower() == 'nan'


def get_cell_value(cell):
    """
    Récupère la valeur d'une cellule, retourne une chaîne vide si la cellule est vide ou None
    """
    if cell is None:
        return ''
    return str(cell.value).strip() if cell.value is not None else ''


def process_excel_file(excel_path):
    """
    Lit le fichier Excel et génère un QR code pour chaque ligne
    """
    try:
        # Lire le fichier Excel avec openpyxl
        workbook = openpyxl.load_workbook(excel_path, data_only=True)
        sheet = workbook.active
        
        # Trouver les en-têtes de colonnes
        headers = {}
        for col_idx, cell in enumerate(next(sheet.iter_rows(min_row=1, max_row=1)), 1):
            if cell.value:
                headers[cell.value.lower()] = col_idx - 1
        
        # Vérifier que les colonnes nécessaires existent
        required_columns = ['prenom', 'nom']
        missing_columns = [col for col in required_columns if col not in headers]
        
        if missing_columns:
            print(f"Colonnes manquantes dans le fichier Excel: {missing_columns}")
            print(f"Colonnes disponibles: {list(headers.keys())}")
            return
        
        # Compter le nombre de lignes avec des données (en excluant l'en-tête)
        data_rows = list(sheet.iter_rows(min_row=2))
        print(f"Traitement de {len(data_rows)} entrées...")
        
        # Générer un QR code pour chaque ligne
        for index, row in enumerate(data_rows):
            # Récupérer les champs obligatoires
            prenom_idx = headers.get('prenom', -1)
            nom_idx = headers.get('nom', -1)
            
            prenom = get_cell_value(row[prenom_idx]) if prenom_idx >= 0 else ''
            nom = get_cell_value(row[nom_idx]) if nom_idx >= 0 else ''
            
            # Ignorer les lignes sans nom ni prénom
            if not prenom and not nom:
                print(f"Ligne {index + 2} ignorée: nom et prénom manquants")
                continue
            
            # Récupérer les champs optionnels
            profession = get_cell_value(row[headers.get('profession', -1)]) if 'profession' in headers else ''
            societe = get_cell_value(row[headers.get('societe', -1)]) if 'societe' in headers else ''
            mobile = get_cell_value(row[headers.get('mobile', -1)]) if 'mobile' in headers else ''
            pro = get_cell_value(row[headers.get('pro', -1)]) if 'pro' in headers else ''
            email = get_cell_value(row[headers.get('email', -1)]) if 'email' in headers else ''
            adresse = get_cell_value(row[headers.get('adresse', -1)]) if 'adresse' in headers else ''
            site_web = get_cell_value(row[headers.get('site_web', -1)]) if 'site_web' in headers else ''
            
            # Créer la vCard
            vcard_data = create_vcard(prenom, nom, profession, societe, mobile, pro, email, adresse, site_web)
            
            # Générer le nom de fichier
            filename = f"{prenom}_{nom}_{index + 1}"
            filename = "".join(c for c in filename if c.isalnum() or c in (' ', '-', '_')).rstrip()
            
            # Générer le QR code
            filepath = generate_qr_svg(vcard_data, filename)
            print(f"QR code généré: {filepath}")
        
        print("Génération terminée!")
        
    except FileNotFoundError:
        print(f"Fichier Excel non trouvé: {excel_path}")
    except Exception as e:
        print(f"Erreur lors du traitement: {e}")


def main():
    """
    Fonction principale
    """
    print("=== Générateur de QR codes vCard ===")
    
    # Le excel est a la racine "contacts.xlsx"
    excel_path = "contacts.xlsx"
    process_excel_file(excel_path)


if __name__ == "__main__":
    main()

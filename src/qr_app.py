import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import qrcode
import qrcode.image.svg
import os
import sys
import base64
import platform
import subprocess
from pathlib import Path
from PIL import Image, ImageTk
import io
import openpyxl
import openpyxl.styles

# Importer les fonctions du générateur
from qr_generator import create_vcard, parse_address, generate_qr_svg, process_excel_file

class QRCodeGeneratorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Générateur de QR Code - Sudalys Services")
        self.root.geometry("800x700")
        self.root.resizable(True, True)
        
        # Ajouter l'icône à la fenêtre principale
        try:
            # Chemin relatif depuis le dossier src vers assets
            icon_path = os.path.join("..", "assets", "logo_sudalys_services.ico")
            if os.path.exists(icon_path):
                self.root.iconbitmap(icon_path)
        except Exception:
            # Si l'icône n'est pas trouvée, continuer sans erreur
            pass
        
        # Configurer le style
        self.style = ttk.Style()
        self.style.configure("TFrame", background="#f0f0f0")
        self.style.configure("TLabel", background="#f0f0f0", font=("Helvetica", 10))
        self.style.configure("TButton", font=("Helvetica", 10))
        self.style.configure("Header.TLabel", font=("Helvetica", 14, "bold"))
        
        # Créer le cadre principal
        self.main_frame = ttk.Frame(root, padding="20")
        self.main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Logo
        try:
            logo_path = "logo_sudalys_services.jpg"
            logo_img = Image.open(logo_path)
            logo_img = logo_img.resize((200, 100), Image.LANCZOS)
            self.logo_photo = ImageTk.PhotoImage(logo_img)
            
            logo_label = ttk.Label(self.main_frame, image=self.logo_photo, background="#f0f0f0")
            logo_label.grid(row=0, column=0, columnspan=2, pady=(0, 20))
        except Exception as e:
            print(f"Erreur lors du chargement du logo: {e}")
        
        # Titre
        title_label = ttk.Label(self.main_frame, text="Générateur de QR Code vCard", style="Header.TLabel")
        title_label.grid(row=1, column=0, columnspan=2, pady=(0, 20))
        
        # Créer les onglets
        self.tab_control = ttk.Notebook(self.main_frame)
        
        # Onglet 1: Saisie manuelle
        self.tab_manual = ttk.Frame(self.tab_control)
        self.tab_control.add(self.tab_manual, text="Saisie manuelle")
        
        # Onglet 2: Import Excel
        self.tab_excel = ttk.Frame(self.tab_control)
        self.tab_control.add(self.tab_excel, text="Import Excel")
        
        self.tab_control.grid(row=2, column=0, columnspan=2, sticky="nsew")
        
        # Configuration de l'onglet Saisie manuelle
        self.setup_manual_tab()
        
        # Configuration de l'onglet Import Excel
        self.setup_excel_tab()
        
        # Cadre pour l'aperçu du QR code
        self.preview_frame = ttk.LabelFrame(self.main_frame, text="Aperçu du QR Code", padding="10")
        self.preview_frame.grid(row=3, column=0, columnspan=2, pady=20, sticky="nsew")
        
        self.preview_label = ttk.Label(self.preview_frame, background="white")
        self.preview_label.pack(padx=10, pady=10, expand=True)
        
        # Configurer les poids des lignes et colonnes pour le redimensionnement
        self.main_frame.columnconfigure(0, weight=1)
        self.main_frame.columnconfigure(1, weight=1)
        self.main_frame.rowconfigure(2, weight=1)
        self.main_frame.rowconfigure(3, weight=1)
        
        # Variables pour stocker les données
        self.current_qr_data = None
        self.current_qr_filename = None
        
        # Données Excel
        self.excel_data = []
        self.excel_headers = {}
        
        # Variable pour optimiser la génération en temps réel
        self.last_update_time = 0
        self.update_delay = 500  # 500ms de délai pour éviter trop de générations
        
        # Variable pour la fenêtre d'aide
        self.help_window = None
        
        # Vider complètement la mémoire au démarrage
        self.clear_all_data()
    
    def setup_manual_tab(self):
        # Créer un cadre pour les champs de saisie
        form_frame = ttk.Frame(self.tab_manual, padding="10")
        form_frame.pack(fill=tk.BOTH, expand=True)
        
        # Créer les champs de saisie
        fields = [
            ("Prénom*", "prenom"),
            ("Nom*", "nom"),
            ("Profession", "profession"),
            ("Société", "societe"),
            ("Mobile", "mobile"),
            ("Téléphone Pro", "pro"),
            ("Email", "email"),
            ("Adresse", "adresse", "Ex: 123 rue de Paris 75001 PARIS I FRANCE"),
            ("Site Web", "site_web")
        ]
        
        self.manual_entries = {}
        
        for i, (label_text, field_name, *hint) in enumerate(fields):
            label = ttk.Label(form_frame, text=label_text)
            label.grid(row=i, column=0, sticky="w", padx=5, pady=5)
            
            entry = ttk.Entry(form_frame, width=40)
            entry.grid(row=i, column=1, sticky="ew", padx=5, pady=5)
            
            # Ajouter l'événement de saisie en temps réel
            entry.bind('<KeyRelease>', self.on_manual_field_change)
            entry.bind('<FocusOut>', self.on_manual_field_change)
            
            self.manual_entries[field_name] = entry
            
            # Ajouter un texte d'aide pour le champ adresse
            if field_name == "adresse" and hint:
                help_frame = ttk.Frame(form_frame)
                help_frame.grid(row=i, column=2, sticky="w", padx=5, pady=5)
                
                help_text = ttk.Label(help_frame, text=hint[0], foreground="gray", font=("Helvetica", 8))
                help_text.pack(side=tk.LEFT)
                
                help_btn = ttk.Button(help_frame, text="?", width=2, command=self.show_address_help)
                help_btn.pack(side=tk.LEFT, padx=5)
        
        # Boutons
        button_frame = ttk.Frame(form_frame)
        button_frame.grid(row=len(fields), column=0, columnspan=2, pady=15)
        
        generate_btn = ttk.Button(button_frame, text="Générer QR Code", command=self.generate_manual_qr)
        generate_btn.pack(side=tk.LEFT, padx=5)
        
        clear_btn = ttk.Button(button_frame, text="Effacer", command=self.clear_manual_form)
        clear_btn.pack(side=tk.LEFT, padx=5)
        
        clear_all_btn = ttk.Button(button_frame, text="Tout Vider", command=self.clear_all_data)
        clear_all_btn.pack(side=tk.LEFT, padx=5)
        
        save_btn = ttk.Button(button_frame, text="Enregistrer QR Code", command=self.save_current_qr)
        save_btn.pack(side=tk.LEFT, padx=5)
        
        # Configurer le redimensionnement
        form_frame.columnconfigure(1, weight=1)
    
    def setup_excel_tab(self):
        # Créer un cadre pour l'import Excel
        excel_frame = ttk.Frame(self.tab_excel, padding="10")
        excel_frame.pack(fill=tk.BOTH, expand=True)
        
        # Champ pour sélectionner le fichier Excel
        file_frame = ttk.Frame(excel_frame)
        file_frame.pack(fill=tk.X, pady=10)
        
        file_label = ttk.Label(file_frame, text="Fichier Excel:")
        file_label.pack(side=tk.LEFT, padx=5)
        
        self.excel_path_var = tk.StringVar()
        self.excel_path_var.set("contacts.xlsx")  # Valeur par défaut
        
        file_entry = ttk.Entry(file_frame, textvariable=self.excel_path_var, width=40)
        file_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        browse_btn = ttk.Button(file_frame, text="Parcourir", command=self.browse_excel)
        browse_btn.pack(side=tk.LEFT, padx=5)
        
        # Boutons Excel
        excel_buttons_frame = ttk.Frame(excel_frame)
        excel_buttons_frame.pack(pady=10)
        
        generate_all_btn = ttk.Button(excel_buttons_frame, text="Générer tous les QR Codes", command=self.generate_all_qr)
        generate_all_btn.pack(side=tk.LEFT, padx=5)
        
        clear_excel_btn = ttk.Button(excel_buttons_frame, text="Vider Excel", command=self.clear_excel_data)
        clear_excel_btn.pack(side=tk.LEFT, padx=5)
        
        template_btn = ttk.Button(excel_buttons_frame, text="Télécharger Template", command=self.download_excel_template)
        template_btn.pack(side=tk.LEFT, padx=5)
        
        # Tableau pour afficher les données Excel
        self.excel_tree = ttk.Treeview(excel_frame, columns=("prenom", "nom", "profession", "societe", "email"), show="headings")
        self.excel_tree.heading("prenom", text="Prénom")
        self.excel_tree.heading("nom", text="Nom")
        self.excel_tree.heading("profession", text="Profession")
        self.excel_tree.heading("societe", text="Société")
        self.excel_tree.heading("email", text="Email")
        
        self.excel_tree.column("prenom", width=100)
        self.excel_tree.column("nom", width=100)
        self.excel_tree.column("profession", width=120)
        self.excel_tree.column("societe", width=120)
        self.excel_tree.column("email", width=150)
        
        scrollbar = ttk.Scrollbar(excel_frame, orient=tk.VERTICAL, command=self.excel_tree.yview)
        self.excel_tree.configure(yscrollcommand=scrollbar.set)
        
        self.excel_tree.pack(fill=tk.BOTH, expand=True, pady=10)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Sélection d'un élément dans le tableau
        self.excel_tree.bind("<<TreeviewSelect>>", self.on_excel_row_select)
        
        # Charger le fichier Excel par défaut s'il existe
        self.load_excel_file()
    
    def browse_excel(self):
        file_path = filedialog.askopenfilename(
            title="Sélectionner un fichier Excel",
            filetypes=[("Fichiers Excel", "*.xlsx;*.xls")]
        )
        if file_path:
            self.excel_path_var.set(file_path)
            self.load_excel_file()
    
    def load_excel_file(self):
        try:
            file_path = self.excel_path_var.get()
            if not os.path.exists(file_path):
                return
                
            # Lire le fichier Excel avec openpyxl
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            sheet = workbook.active
            
            # Trouver les en-têtes de colonnes
            self.excel_headers = {}
            header_row = next(sheet.iter_rows(min_row=1, max_row=1))
            for col_idx, cell in enumerate(header_row):
                if cell.value:
                    self.excel_headers[cell.value.lower()] = col_idx
            
            # Lire toutes les données (seulement les lignes non vides)
            self.excel_data = []
            for row in sheet.iter_rows(min_row=2):
                row_data = {}
                has_data = False
                
                for header, col_idx in self.excel_headers.items():
                    cell = row[col_idx]
                    value = str(cell.value).strip() if cell.value is not None else ''
                    row_data[header] = value
                    if value:  # Si au moins une cellule a des données
                        has_data = True
                
                # N'ajouter que les lignes qui ont au moins une donnée
                if has_data:
                    self.excel_data.append(row_data)
            
            # Effacer les données existantes dans le tableau
            for item in self.excel_tree.get_children():
                self.excel_tree.delete(item)
            
            # Remplir le tableau avec les données
            for index, row_data in enumerate(self.excel_data):
                values = (
                    row_data.get('prenom', ''),
                    row_data.get('nom', ''),
                    row_data.get('profession', ''),
                    row_data.get('societe', ''),
                    row_data.get('email', '')
                )
                self.excel_tree.insert('', tk.END, values=values, iid=str(index))
            
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors du chargement du fichier Excel: {e}")
    
    def on_excel_row_select(self, event):
        selected_items = self.excel_tree.selection()
        if selected_items:
            try:
                index = int(selected_items[0])
                # Vérifier que l'index est valide
                if 0 <= index < len(self.excel_data):
                    row_data = self.excel_data[index]
                    
                    # Vérifier que la ligne a des données valides
                    prenom = row_data.get('prenom', '').strip()
                    nom = row_data.get('nom', '').strip()
                    
                    if prenom or nom:  # Au moins un des deux champs requis
                        # Générer le QR code pour cette ligne
                        self.generate_qr_from_row(row_data, index)
                    else:
                        # Désélectionner silencieusement la ligne vide
                        self.excel_tree.selection_remove(selected_items[0])
                else:
                    # Index invalide, désélectionner silencieusement
                    self.excel_tree.selection_remove(selected_items[0])
            except (ValueError, IndexError):
                # Erreur de conversion ou d'index, désélectionner silencieusement
                if selected_items:
                    self.excel_tree.selection_remove(selected_items[0])
    
    def generate_qr_from_row(self, row_data, index):
        try:
            # Récupérer les données
            prenom = row_data.get('prenom', '')
            nom = row_data.get('nom', '')
            
            if not prenom and not nom:
                messagebox.showwarning("Données incomplètes", "Le prénom et le nom sont requis.")
                return
            
            profession = row_data.get('profession', '')
            societe = row_data.get('societe', '')
            mobile = row_data.get('mobile', '')
            pro = row_data.get('pro', '')
            email = row_data.get('email', '')
            adresse = row_data.get('adresse', '')
            site_web = row_data.get('site_web', '')
            
            # Créer la vCard
            vcard_data = create_vcard(prenom, nom, profession, societe, mobile, pro, email, adresse, site_web)
            
            # Générer le nom de fichier
            filename = f"{prenom}_{nom}_{index + 1}"
            filename = "".join(c for c in filename if c.isalnum() or c in (' ', '-', '_')).rstrip()
            
            # Stocker les données actuelles
            self.current_qr_data = vcard_data
            self.current_qr_filename = filename
            
            # Afficher l'aperçu
            self.update_qr_preview(vcard_data)
            
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de la génération du QR code: {e}")
    
    def generate_manual_qr(self):
        try:
            # Récupérer les données des champs
            prenom = self.manual_entries["prenom"].get().strip()
            nom = self.manual_entries["nom"].get().strip()
            
            if not prenom or not nom:
                messagebox.showwarning("Données incomplètes", "Le prénom et le nom sont requis.")
                return
            
            profession = self.manual_entries["profession"].get().strip()
            societe = self.manual_entries["societe"].get().strip()
            mobile = self.manual_entries["mobile"].get().strip()
            pro = self.manual_entries["pro"].get().strip()
            email = self.manual_entries["email"].get().strip()
            adresse = self.manual_entries["adresse"].get().strip()
            site_web = self.manual_entries["site_web"].get().strip()
            
            # Créer la vCard
            vcard_data = create_vcard(prenom, nom, profession, societe, mobile, pro, email, adresse, site_web)
            
            # Générer le nom de fichier
            filename = f"{prenom}_{nom}_manual"
            filename = "".join(c for c in filename if c.isalnum() or c in (' ', '-', '_')).rstrip()
            
            # Stocker les données actuelles
            self.current_qr_data = vcard_data
            self.current_qr_filename = filename
            
            # Afficher l'aperçu
            self.update_qr_preview(vcard_data)
            
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de la génération du QR code: {e}")
    
    def update_qr_preview(self, data):
        try:
            # Générer le QR code en mémoire
            qr = qrcode.QRCode(
                version=1,
                error_correction=qrcode.constants.ERROR_CORRECT_L,
                box_size=10,
                border=4,
            )
            qr.add_data(data)
            qr.make(fit=True)
            
            img = qr.make_image(fill_color="black", back_color="white")
            
            # Convertir l'image pour l'affichage Tkinter
            img_buffer = io.BytesIO()
            img.save(img_buffer, format="PNG")
            img_buffer.seek(0)
            
            tk_img = Image.open(img_buffer)
            tk_img = tk_img.resize((200, 200), Image.LANCZOS)
            self.preview_photo = ImageTk.PhotoImage(tk_img)
            
            # Mettre à jour l'aperçu
            self.preview_label.configure(image=self.preview_photo)
            
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de la génération de l'aperçu: {e}")
    
    def save_current_qr(self):
        if not self.current_qr_data or not self.current_qr_filename:
            messagebox.showwarning("Aucun QR code", "Veuillez d'abord générer un QR code.")
            return
        
        try:
            # Générer le QR code SVG
            filepath = generate_qr_svg(self.current_qr_data, self.current_qr_filename)
            
            # Déterminer si nous sommes en mode compilé ou développement
            if getattr(sys, 'frozen', False):
                # Mode compilé
                directory = filepath.parent
                messagebox.showinfo("Succès", f"QR code enregistré avec succès dans le dossier:\n{directory}\n\nNom du fichier: {filepath.name}")
            else:
                # Mode développement
                messagebox.showinfo("Succès", f"QR code enregistré avec succès: {filepath}")
                
            # Proposer d'ouvrir le dossier contenant le QR code
            if messagebox.askyesno("Ouvrir le dossier", "Voulez-vous ouvrir le dossier contenant le QR code?"):
                if platform.system() == "Windows":
                    os.startfile(filepath.parent)
                elif platform.system() == "Darwin":  # macOS
                    subprocess.Popen(["open", filepath.parent])
                else:  # Linux
                    subprocess.Popen(["xdg-open", filepath.parent])
                    
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de l'enregistrement du QR code: {e}")
    
    def generate_all_qr(self):
        try:
            file_path = self.excel_path_var.get()
            if not os.path.exists(file_path):
                messagebox.showwarning("Fichier introuvable", f"Le fichier {file_path} n'existe pas.")
                return
            
            # Utiliser la fonction existante pour traiter le fichier Excel
            process_excel_file(file_path)
            
            # Déterminer le dossier de sortie en fonction du mode (compilé ou développement)
            
            if getattr(sys, 'frozen', False):
                # Mode compilé
                if platform.system() == 'Windows':
                    user_docs = os.path.join(os.path.expanduser('~'), 'Documents')
                    output_dir = Path(user_docs) / "Sudalys_QR_Codes"
                else:
                    output_dir = Path(os.path.expanduser('~')) / "Sudalys_QR_Codes"
            else:
                # Mode développement
                output_dir = Path("qr_codes")
            
            messagebox.showinfo("Succès", f"Tous les QR codes ont été générés avec succès dans le dossier:\n{output_dir}")
            
            # Proposer d'ouvrir le dossier contenant les QR codes
            if messagebox.askyesno("Ouvrir le dossier", "Voulez-vous ouvrir le dossier contenant les QR codes?"):
                if platform.system() == "Windows":
                    os.startfile(output_dir)
                elif platform.system() == "Darwin":  # macOS
                    subprocess.Popen(["open", output_dir])
                else:  # Linux
                    subprocess.Popen(["xdg-open", output_dir])
            
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de la génération des QR codes: {e}")
    
    def clear_manual_form(self):
        """Vide uniquement les champs de saisie manuelle"""
        for entry in self.manual_entries.values():
            entry.delete(0, tk.END)
    
    def clear_excel_data(self):
        """Vide les données Excel et remet le chemin par défaut"""
        # Vider le tableau Excel
        for item in self.excel_tree.get_children():
            self.excel_tree.delete(item)
        
        # Vider les données en mémoire
        self.excel_data = []
        self.excel_headers = {}
        
        # Remettre le chemin par défaut
        self.excel_path_var.set("contacts.xlsx")
        
    
    def clear_all_data(self):
        """Vide complètement toutes les données de l'application"""
        # Vider les champs de saisie manuelle
        if hasattr(self, 'manual_entries'):
            for entry in self.manual_entries.values():
                entry.delete(0, tk.END)
        
        # Vider les données Excel
        if hasattr(self, 'excel_tree'):
            for item in self.excel_tree.get_children():
                self.excel_tree.delete(item)
        
        # Réinitialiser toutes les variables de données
        self.excel_data = []
        self.excel_headers = {}
        self.current_qr_data = None
        self.current_qr_filename = None
        
        # Remettre le chemin Excel par défaut
        if hasattr(self, 'excel_path_var'):
            self.excel_path_var.set("contacts.xlsx")
        
        # Vider l'aperçu du QR code
        if hasattr(self, 'preview_label'):
            self.preview_label.configure(image="")
            if hasattr(self, 'preview_photo'):
                del self.preview_photo
        
        # Pas de message de confirmation pour éviter les popups
    
    def on_manual_field_change(self, event=None):
        """Génère automatiquement le QR code quand les champs changent"""
        import time
        current_time = time.time() * 1000  # Convertir en millisecondes
        
        # Optimisation : éviter trop de générations rapprochées
        if current_time - self.last_update_time < self.update_delay:
            # Programmer une génération différée
            self.root.after(self.update_delay, self.delayed_qr_generation)
            return
        
        self.last_update_time = current_time
        self.generate_realtime_qr()
    
    def delayed_qr_generation(self):
        """Génération différée pour optimiser les performances"""
        import time
        current_time = time.time() * 1000
        
        # Vérifier si une nouvelle saisie n'a pas eu lieu entre temps
        if current_time - self.last_update_time >= self.update_delay:
            self.generate_realtime_qr()
    
    def generate_realtime_qr(self):
        """Génère le QR code en temps réel pendant la saisie"""
        try:
            # Récupérer les données des champs
            prenom = self.manual_entries["prenom"].get().strip()
            nom = self.manual_entries["nom"].get().strip()
            
            # Générer seulement s'il y a au moins un prénom ou nom
            if prenom or nom:
                profession = self.manual_entries["profession"].get().strip()
                societe = self.manual_entries["societe"].get().strip()
                mobile = self.manual_entries["mobile"].get().strip()
                pro = self.manual_entries["pro"].get().strip()
                email = self.manual_entries["email"].get().strip()
                adresse = self.manual_entries["adresse"].get().strip()
                site_web = self.manual_entries["site_web"].get().strip()
                
                # Créer la vCard
                vcard_data = create_vcard(prenom, nom, profession, societe, mobile, pro, email, adresse, site_web)
                
                # Générer le nom de fichier
                filename = f"{prenom}_{nom}_realtime"
                filename = "".join(c for c in filename if c.isalnum() or c in (' ', '-', '_')).rstrip()
                
                # Stocker les données actuelles
                self.current_qr_data = vcard_data
                self.current_qr_filename = filename
                
                # Afficher l'aperçu
                self.update_qr_preview(vcard_data)
            else:
                # Vider l'aperçu si pas de données
                if hasattr(self, 'preview_label'):
                    self.preview_label.configure(image="")
                    if hasattr(self, 'preview_photo'):
                        del self.preview_photo
                
        except Exception as e:
            # En cas d'erreur, ne pas afficher de message pour ne pas gêner la saisie
            pass
    
    def download_excel_template(self):
        """Crée et télécharge un template Excel avec les colonnes attendues"""
        try:
            # Créer un nouveau classeur Excel
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Contacts"
            
            # Définir les en-têtes de colonnes
            headers = [
                "Prenom", "Nom", "Profession", "Societe", "Mobile", 
                "Pro", "Email", "Adresse", "Site_Web"
            ]
            
            # Ajouter les en-têtes
            for col_idx, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col_idx)
                cell.value = header
                cell.font = openpyxl.styles.Font(bold=True)
                cell.fill = openpyxl.styles.PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
            
            # Ajouter quelques exemples de données
            examples = [
                ["Jean", "DUPONT", "Développeur", "SUDALYS", "06.00.00.00.00", "01.00.00.00.00", "jean.dupont@sudalys.fr", "123 rue de Paris 75001 PARIS", "www.sudalys.fr"],
                ["Marie", "MARTIN", "Chef de projet", "SUDALYS", "06.00.00.00.00", "01.00.00.00.00", "marie.martin@sudalys.fr", "456 avenue des Champs 75008 PARIS", ""],
                ["Pierre", "BERNARD", "Commercial", "SUDALYS", "06.00.00.00.00", "", "pierre.bernard@sudalys.fr", "", ""]
            ]
            
            for row_idx, example in enumerate(examples, 2):
                for col_idx, value in enumerate(example, 1):
                    sheet.cell(row=row_idx, column=col_idx, value=value)
            
            # Ajuster la largeur des colonnes
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                sheet.column_dimensions[column_letter].width = adjusted_width
            
            # Demander où sauvegarder le fichier
            file_path = filedialog.asksaveasfilename(
                title="Enregistrer le template Excel",
                defaultextension=".xlsx",
                filetypes=[("Fichiers Excel", "*.xlsx")]
            )
            
            if file_path:
                # Sauvegarder le fichier
                workbook.save(file_path)
                messagebox.showinfo("Template créé", f"Template Excel créé avec succès :\n{file_path}\n\nLe fichier contient :\n- Les colonnes attendues par l'application\n- Quelques exemples de données\n- Un formatage professionnel")
                
                # Proposer d'ouvrir le fichier
                if messagebox.askyesno("Ouvrir le fichier", "Voulez-vous ouvrir le template Excel ?"):
                    if platform.system() == "Windows":
                        os.startfile(file_path)
                    elif platform.system() == "Darwin":  # macOS
                        subprocess.Popen(["open", file_path])
                    else:  # Linux
                        subprocess.Popen(["xdg-open", file_path])
            
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de la création du template : {e}")
            
    def show_address_help(self):
        # Vérifier si la fenêtre d'aide existe déjà et est visible
        if self.help_window and self.help_window.winfo_exists():
            # Ramener la fenêtre au premier plan
            self.help_window.lift()
            self.help_window.focus_force()
            return
        
        help_text = """Formats d'adresse pris en charge:

1. Format standard: 123 rue de Paris 75001 PARIS
2. Avec pays: 123 rue de Paris 75001 PARIS I FRANCE
3. Format inversé: 123 rue de Paris PARIS 75001
4. Format avec parenthèses: 123 rue de Paris PARIS (75001)
5. Format international: 123 rue de Paris F-75001 PARIS

Le système essaiera d'extraire automatiquement:
- La rue
- La ville
- Le code postal
- Le pays (si spécifié)

Pour de meilleurs résultats, utilisez le format standard."""
        
        self.help_window = tk.Toplevel(self.root)
        self.help_window.title("Aide - Format d'adresse")
        self.help_window.geometry("500x350")
        
        # Ajouter l'icône à la fenêtre d'aide
        try:
            # Chemin relatif depuis le dossier src vers assets
            icon_path = os.path.join("..", "assets", "logo_sudalys_services.ico")
            if os.path.exists(icon_path):
                self.help_window.iconbitmap(icon_path)
        except Exception:
            # Si l'icône n'est pas trouvée, continuer sans erreur
            pass
        
        # Fonction pour nettoyer la référence quand la fenêtre se ferme
        def on_help_window_close():
            self.help_window = None
        
        self.help_window.protocol("WM_DELETE_WINDOW", lambda: [self.help_window.destroy(), on_help_window_close()])
        
        text_widget = tk.Text(self.help_window, wrap=tk.WORD, padx=10, pady=10)
        text_widget.insert(tk.END, help_text)
        text_widget.config(state=tk.DISABLED)
        text_widget.pack(fill=tk.BOTH, expand=True)
        
        close_btn = ttk.Button(self.help_window, text="Fermer", command=lambda: [self.help_window.destroy(), on_help_window_close()])
        close_btn.pack(pady=10)

def main():
    root = tk.Tk()
    app = QRCodeGeneratorApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()

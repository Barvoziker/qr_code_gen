import qrcode
import qrcode.image.svg
import os
import sys
import base64
import platform
from pathlib import Path
import openpyxl

def parse_address(adresse_complete):
    """
    Parse une adresse et la convertit au format ADR vCard
    Format ADR: Boîte postale;Adresse étendue;Nom de rue;Ville;Région;Code postal;Pays
    """
    if not adresse_complete or str(adresse_complete).strip().lower() == 'nan':
        return ";;;;;;"
        
    try:
        adresse_complete = str(adresse_complete).strip()
        
        # Initialiser les variables
        rue = ''
        ville = ''
        code_postal = ''
        pays = ''
        region = ''
        
        # Étape 1: Extraire le pays s'il est présent
        import re
        
        # Rechercher le pays à la fin (après I, FR, FRANCE, etc.)
        pays_patterns = [
            r'\s+I\s+([A-Z][A-Za-z\s]+)$',  # Format " I FRANCE"
            r'\s+FR\s+([A-Z][A-Za-z\s]+)$',  # Format " FR FRANCE"
            r'\s+FRANCE$',                   # Juste "FRANCE" à la fin
            r'\s+([A-Z]{2,3})$'              # Code pays à la fin (FR, USA, etc.)
        ]
        
        for pattern in pays_patterns:
            pays_match = re.search(pattern, adresse_complete)
            if pays_match:
                if len(pays_match.groups()) > 0:
                    pays = pays_match.group(1).strip()
                else:
                    pays = 'FRANCE'  # Si on trouve juste le mot FRANCE
                adresse_complete = adresse_complete[:pays_match.start()].strip()
                break
        
        # Si aucun pays trouvé mais contient "FRANCE" quelque part
        if not pays and 'FRANCE' in adresse_complete.upper():
            pays = 'FRANCE'
            adresse_complete = adresse_complete.upper().replace('FRANCE', '').strip()
        
        # Étape 2: Extraire le code postal et la ville
        # Différents formats possibles: "75001 PARIS", "F-75001 PARIS", "PARIS (75001)", etc.
        cp_ville_patterns = [
            r'(\d{5})\s+([^\d]+)$',               # Format "75001 PARIS"
            r'F-(\d{5})\s+([^\d]+)$',             # Format "F-75001 PARIS"
            r'([^\d\(]+)\s*\((\d{5})\)$',         # Format "PARIS (75001)"
            r'([^\d\(]+)\s+(\d{5})$',             # Format "PARIS 75001"
        ]
        
        for pattern in cp_ville_patterns:
            cp_ville_match = re.search(pattern, adresse_complete)
            if cp_ville_match:
                # Selon le format, l'ordre des groupes peut varier
                if pattern == r'([^\d\(]+)\s*\((\d{5})\)$' or pattern == r'([^\d\(]+)\s+(\d{5})$':
                    ville = cp_ville_match.group(1).strip()
                    code_postal = cp_ville_match.group(2).strip()
                else:
                    code_postal = cp_ville_match.group(1).strip()
                    ville = cp_ville_match.group(2).strip()
                
                # Le reste est l'adresse de rue
                rue = adresse_complete[:cp_ville_match.start()].strip()
                break
        
        # Si pas de match, essayer de trouver juste un code postal (5 chiffres)
        if not code_postal:
            cp_match = re.search(r'(\d{5})', adresse_complete)
            if cp_match:
                code_postal = cp_match.group(1)
                # Essayer de trouver la ville après le code postal
                reste = adresse_complete[cp_match.end():].strip()
                if reste:
                    # Prendre le premier mot après le code postal comme ville
                    ville_match = re.match(r'([A-Za-z\s\-]+)', reste)
                    if ville_match:
                        ville = ville_match.group(1).strip()
                        # Le reste avant le code postal est la rue
                        rue = adresse_complete[:cp_match.start()].strip()
                    else:
                        # Pas de ville trouvée, tout ce qui reste est la rue
                        rue = adresse_complete.strip()
                else:
                    # Pas de texte après le code postal, tout ce qui précède est la rue
                    rue = adresse_complete[:cp_match.start()].strip()
            else:
                # Pas de code postal trouvé, tout est considéré comme rue
                rue = adresse_complete.strip()
        
        # Format ADR: Boîte postale;Adresse étendue;Nom de rue;Ville;Région;Code postal;Pays
        return f";;{rue};{ville};{region};{code_postal};{pays}"
    
    except Exception as e:
        # En cas d'erreur, retourner l'adresse dans le champ rue
        print(f"Erreur lors du parsing de l'adresse: {e}")
        return f";;{adresse_complete};;;;"


def create_vcard(prenom, nom, profession, societe, mobile, pro, email, adresse, site_web):
    """
    Crée une vCard au format texte avec tous les champs
    """
    # Construire le nom complet
    full_name = f"{prenom} {nom}".strip()
    
    # Commencer la vCard
    vcard_lines = [
        "BEGIN:VCARD",
        "VERSION:3.0",
        f"FN:{full_name}",
        f"N:{nom};{prenom};;;"
    ]
    
    # Ajouter les champs optionnels s'ils existent
    if profession and str(profession).strip() and str(profession).strip().lower() != 'nan':
        vcard_lines.append(f"TITLE:{profession}")
    
    if societe and str(societe).strip() and str(societe).strip().lower() != 'nan':
        vcard_lines.append(f"ORG:{societe}")
    
    if email and str(email).strip() and str(email).strip().lower() != 'nan':
        vcard_lines.append(f"EMAIL:{email}")
    
    if mobile and str(mobile).strip() and str(mobile).strip().lower() != 'nan':
        vcard_lines.append(f"TEL;TYPE=CELL:{mobile}")
    
    if pro and str(pro).strip() and str(pro).strip().lower() != 'nan':
        vcard_lines.append(f"TEL;TYPE=WORK:{pro}")
    
    if adresse and str(adresse).strip() and str(adresse).strip().lower() != 'nan':
        parsed_address = parse_address(str(adresse).strip())
        vcard_lines.append(f"ADR:{parsed_address}")
    
    if site_web and str(site_web).strip() and str(site_web).strip().lower() != 'nan':
        # Ajouter http:// si pas de protocole
        url = str(site_web).strip()
        if not url.startswith(('http://', 'https://')):
            url = f"https://{url}"
        vcard_lines.append(f"URL:{url}")
    
    vcard_lines.append("END:VCARD")
    
    return "\n".join(vcard_lines)


def generate_qr_svg(data, filename):
    """
    Génère un QR code en format SVG
    """
    factory = qrcode.image.svg.SvgPathImage
    img = qrcode.make(data, image_factory=factory)
    
    # Déterminer le chemin approprié pour les QR codes
    # Si l'application est compilée en .exe, utiliser le dossier de l'utilisateur
    if getattr(sys, 'frozen', False):
        # L'application est compilée avec PyInstaller
        # Utiliser le dossier Documents de l'utilisateur
        
        if platform.system() == 'Windows':
            # Sur Windows, utiliser le dossier Documents
            user_docs = os.path.join(os.path.expanduser('~'), 'Documents')
            output_dir = Path(user_docs) / "Sudalys_QR_Codes"
        else:
            # Sur d'autres systèmes, utiliser le dossier home
            output_dir = Path(os.path.expanduser('~')) / "Sudalys_QR_Codes"
    else:
        # En mode développement, utiliser le dossier local
        output_dir = Path("qr_codes")
    
    # Créer le dossier de sortie s'il n'existe pas
    output_dir.mkdir(exist_ok=True, parents=True)
    
    # Sauvegarder le fichier SVG
    filepath = output_dir / f"{filename}.svg"
    with open(filepath, 'wb') as f:
        img.save(f)
    
    return filepath


def is_empty_cell(cell):
    """
    Vérifie si une cellule est vide ou contient 'nan'
    """
    if cell is None:
        return True
    cell_value = str(cell).strip()
    return cell_value == '' or cell_value.lower() == 'nan'


def get_cell_value(cell):
    """
    Récupère la valeur d'une cellule, retourne une chaîne vide si la cellule est vide ou None
    """
    if cell is None:
        return ''
    return str(cell.value).strip() if cell.value is not None else ''


def process_excel_file(excel_path):
    """
    Lit le fichier Excel et génère un QR code pour chaque ligne
    """
    try:
        # Lire le fichier Excel avec openpyxl
        workbook = openpyxl.load_workbook(excel_path, data_only=True)
        sheet = workbook.active
        
        # Trouver les en-têtes de colonnes
        headers = {}
        for col_idx, cell in enumerate(next(sheet.iter_rows(min_row=1, max_row=1)), 1):
            if cell.value:
                headers[cell.value.lower()] = col_idx - 1
        
        # Vérifier que les colonnes nécessaires existent
        required_columns = ['prenom', 'nom']
        missing_columns = [col for col in required_columns if col not in headers]
        
        if missing_columns:
            print(f"Colonnes manquantes dans le fichier Excel: {missing_columns}")
            print(f"Colonnes disponibles: {list(headers.keys())}")
            return
        
        # Compter le nombre de lignes avec des données (en excluant l'en-tête)
        data_rows = list(sheet.iter_rows(min_row=2))
        print(f"Traitement de {len(data_rows)} entrées...")
        
        # Générer un QR code pour chaque ligne
        for index, row in enumerate(data_rows):
            # Récupérer les champs obligatoires
            prenom_idx = headers.get('prenom', -1)
            nom_idx = headers.get('nom', -1)
            
            prenom = get_cell_value(row[prenom_idx]) if prenom_idx >= 0 else ''
            nom = get_cell_value(row[nom_idx]) if nom_idx >= 0 else ''
            
            # Ignorer les lignes sans nom ni prénom
            if not prenom and not nom:
                print(f"Ligne {index + 2} ignorée: nom et prénom manquants")
                continue
            
            # Récupérer les champs optionnels
            profession = get_cell_value(row[headers.get('profession', -1)]) if 'profession' in headers else ''
            societe = get_cell_value(row[headers.get('societe', -1)]) if 'societe' in headers else ''
            mobile = get_cell_value(row[headers.get('mobile', -1)]) if 'mobile' in headers else ''
            pro = get_cell_value(row[headers.get('pro', -1)]) if 'pro' in headers else ''
            email = get_cell_value(row[headers.get('email', -1)]) if 'email' in headers else ''
            adresse = get_cell_value(row[headers.get('adresse', -1)]) if 'adresse' in headers else ''
            site_web = get_cell_value(row[headers.get('site_web', -1)]) if 'site_web' in headers else ''
            
            # Créer la vCard
            vcard_data = create_vcard(prenom, nom, profession, societe, mobile, pro, email, adresse, site_web)
            
            # Générer le nom de fichier
            filename = f"{prenom}_{nom}_{index + 1}"
            filename = "".join(c for c in filename if c.isalnum() or c in (' ', '-', '_')).rstrip()
            
            # Générer le QR code
            filepath = generate_qr_svg(vcard_data, filename)
            print(f"QR code généré: {filepath}")
        
        print("Génération terminée!")
        
    except FileNotFoundError:
        print(f"Fichier Excel non trouvé: {excel_path}")
    except Exception as e:
        print(f"Erreur lors du traitement: {e}")


def main():
    """
    Fonction principale
    """
    print("=== Générateur de QR codes vCard ===")
    
    # Le excel est a la racine "contacts.xlsx"
    excel_path = "contacts.xlsx"
    process_excel_file(excel_path)


if __name__ == "__main__":
    main()
